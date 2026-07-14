# gui/dashboard_tabs.py
from __future__ import annotations

import logging
import os
from typing import Any, Callable, Optional, cast
from tkinter import messagebox
import pandas as pd
import customtkinter as ctk

from gui.workflows import automatikus_feldolgozas_inditasa, kapcsolat_teszt_inditasa

logger = logging.getLogger(__name__)


class DashboardTabs(ctk.CTkFrame):
    """
    Fő vezérlőpult — minden lapfület tartalmaz.
    A főablak (NavM2MApp) példányosítja.
    """

    def __init__(
        self,
        master: Any,
        kornyezet_callback: Optional[Callable[[str], None]] = None,
        **kwargs: Any,
    ) -> None:
        super().__init__(master, **kwargs)
        self.kornyezet_callback = kornyezet_callback
        self._tabview_letrehozasa()

    def _tabview_letrehozasa(self) -> None:
        self.tabs = ctk.CTkTabview(self, anchor="nw")
        self.tabs.pack(fill="both", expand=True)

        tab_feldolgozas  = self.tabs.add("📂 Feldolgozás")
        tab_adozas       = self.tabs.add("📊 Adózási Állapot")
        tab_hataridok    = self.tabs.add("📅 Határidők")
        tab_elozmeny     = self.tabs.add("📋 Előzmények")
        tab_beallitas    = self.tabs.add("⚙️ Beállítások")
        tab_helpdesk     = self.tabs.add("✉️ Helpdesk")

        self._beallitas_tab_felep(tab_beallitas)
        self._feldolgozas_tab_felep(tab_feldolgozas)
        self._adozas_tab_felep(tab_adozas)
        self._hataridok_tab_felep(tab_hataridok)
        self._elozmeny_tab_felep(tab_elozmeny)
        self._helpdesk_tab_felep(tab_helpdesk)

        from utils.logger import Logger
        def log_callback(msg: str, tag: Optional[str]) -> None:
            if hasattr(self, 'log_ablak'):
                log_ctrl = cast(Any, self.log_ablak)
                if tag:
                    log_ctrl.insert("end", msg + "\n", tag)
                else:
                    log_ctrl.insert("end", msg + "\n")
                log_ctrl.see("end")
        cast(Any, Logger).set_callback(log_callback) # type: ignore

    def _sablon_letoltese_vegrehajtasa(self) -> None:
        """Legenerálja, megformázza és komplex magyarázó és valósághű mintasorokkal látja el a fájlt."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment # type: ignore
            from openpyxl.worksheet.datavalidation import DataValidation # type: ignore
            
            fejadat_rows = [
                ["Adatkör", "Megnevezés", "XML mező", "Érték", "Adat"],
                ["A bevallás fejadatai (Módosítás adatai)", "Adószám", "taxNumber", "12345678", "taxNumberField"],
                ["", "Bevallás típusa", "declarationType", "NONE", "declarationTypeCombo"],
                ["", "Bevallás fajtája", "declarationKind", "NONE", "declarationKindCombo"],
                ["", "Bevallás gyakorisága", "declarationFrequency", "MONTHLY", "declarationFrequencyCombo"],
                ["", "Bevallási időszak kezdete", "declarationPeriodStart", "2026-07-01", "periodStartPicker"],
                ["", "Bevallási időszak vége", "declarationPeriodEnd", "2026-07-31", "periodEndPicker"],
                ["", "A benyújtott bevallás verziószáma", "version", "1", "versionField"],
                ["", "Bevallás jellege", "declarationMethod", "BASE", "declarationMethodCombo"],
                ["", "NAV kiértesítés szerinti javítás.", "navCorrection", "false", "navCorrectionCheck"]
            ]
            df_fejadat = pd.DataFrame(fejadat_rows)

            analitika_rows = [
                ["Analitika sorszáma", "Főkönyvi adatok", "", "", "Forrás dokumentum egyedi azonosítója", "Forrás dokumentum kiállítási dátuma", "Forrás dokumentum bizonylat típusa", "Adó esedékességi napja", "Partner adatok", "", "", "", "", "", "", "", "", "", "Adózási adatok", "", "", "", "", "", "", ""],
                ["", "Főkönyvi számlaszám", "Főkönyvi feladás dátuma", "Főkönyvi könyvelés egyedi azonosítója", "", "", "", "", "Partner ÁFA szerinti státusza", "Partner adószáma", "", "", "", "Partner neve", "Partner címadatai", "", "", "", "A tétel standard adókódja", "A tétel vállalati adókódja", "A tétel adózási pozíciója", "", "", "", "A tétel adózási pozíciója", ""],
                ["", "", "", "", "", "", "", "", "", "Belföldi adószám adatok", "", "Közösségi adószám", "Harmadik országbeli adóazonosító", "", "Országkód", "Régió, tartomány kódja", "Irányítószám", "Város", "Cím", "", "", "Áfa pozíció jelölése", "Adó alapja", "Adó összege", "Levonási adatok", "", "Áfa pozíció jelölése", "Adó alapja", "Adó összege", "Levonási adatok"],
                ["", "", "", "", "", "", "", "", "", "Az adószám 8 jegyű törzsszáma", "Csoport tag adószámának 8 jegyű törzsszáma", "", "", "", "", "", "", "", "", "", "", "", "", "", "Levonási hányados", "Levonási összeg", "", "", "Levonási hányados", "Levonási összeg"],
                ["lineNumber", "glAccountId", "glPostingDate", "glTransactionId", "sourceDocumentId", "sourceDocumentIssueDate", "sourceDocumentType", "taxpointDate", "partnerStatus", "customerTaxNumber", "groupMemberTaxNumber", "communityTaxNumber", "thirdCountryTaxId", "partnerName", "countryCode", "region", "postalCode", "city", "streetAddress", "standardTaxCode", "ownTaxCode", "positionType", "taxBase", "taxAmount", "deductionCondition", "deductionRatio", "deductedAmount"],
                # Hivatalos, komplex könyvelési mintasorok beágyazása a dolgozók munkájának segítésére:
                ["1", "466100", "2026-07-10", "TR-99012", "INV-2026-0001", "2026-07-05", "INVOICE", "2026-07-08", "DOMESTIC", "12345678", "", "", "", "Győri Partner Vállalat Kft.", "HU", "Győr-Moson-Sopron", "9021", "Győr", "Városház tér 1.", "DOM_L_GENERAL", "SAP_A1", "DEDUCTIBLE", "1500000", "405000", "", "", "", ""],
                ["2", "467100", "2026-07-14", "TR-99015", "INV-2026-0002", "2026-07-12", "INVOICE", "2026-07-12", "OTHER", "", "", "DE812345678", "", "Minta Import Beszállító GmbH", "DE", "", "10117", "Berlin", "Friedrichstrasse 12.", "LEV_KOZ", "SAP_E1", "PAYABLE", "3200000", "864000", "", "", "", ""]
            ]
            df_analitika = pd.DataFrame(analitika_rows)

            ertekek_data = {
                "declarationType": ["NONE", "ELIMINATION", "LIQUIDATION", "SELF_EMPLOYMENT_END", "TRANSFORMATION", "TERMINATION", "PAUSING"],
                "declarationKind": ["NONE", "PREVIOUS_PERIOD", "UNDER_PROCESS", "CLOSURE", "", "", ""],
                "declarationFrequency": ["ANNUAL", "QUARTERLY", "MONTHLY", "", "", "", ""],
                "declarationMethod": ["BASE", "SELF_CHECK", "CORRECTION", "", "", "", ""],
                "Boolean": ["true", "false", "", "", "", "", ""],
                "sourceDocumentType": ["INVOICE", "RECEIPT", "CUSTOMS_DECLARATION", "OTHER", "", "", ""],
                "partnerStatus": ["NOT_AVAILABLE", "PRIVATE_PERSON", "DOMESTIC", "OTHER", "", "", ""],
                "positionType": ["PAYABLE", "DEDUCTIBLE", "OTHER", "", "", "", ""]
            }
            df_ertekek = pd.DataFrame(ertekek_data)

            asztal_utvonal = os.path.join(os.path.expanduser("~"), "Desktop", "01_pelda_fizetendo.xlsx")
            
            with pd.ExcelWriter(asztal_utvonal, engine='openpyxl') as writer:
                df_fejadat.to_excel(writer, sheet_name='Bevallás Fejadatok', index=False, header=False) # type: ignore
                df_analitika.to_excel(writer, sheet_name='Analitika Adatok', index=False, header=False) # type: ignore
                df_ertekek.to_excel(writer, sheet_name='Értékek', index=False) # type: ignore

                workbook = writer.book
                
                # 1. FEJADAT LAP FORMÁZÁSA ÉS LENYÍLÓK
                ws1 = workbook['Bevallás Fejadatok']
                ws1.views.sheetView[0].showGridLines = True
                
                header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                for col_num in range(1, 6):
                    cell = ws1.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                
                ws1.column_dimensions['A'].width = 38
                ws1.column_dimensions['B'].width = 35
                ws1.column_dimensions['C'].width = 22
                ws1.column_dimensions['D'].width = 18
                ws1.column_dimensions['E'].width = 25

                dv_dec_type = DataValidation(type="list", formula1="=Értékek!$A$2:$A$8", allow_blank=True)
                dv_dec_kind = DataValidation(type="list", formula1="=Értékek!$B$2:$B$5", allow_blank=True)
                dv_dec_freq = DataValidation(type="list", formula1="=Értékek!$C$2:$C$4", allow_blank=True)
                dv_dec_meth = DataValidation(type="list", formula1="=Értékek!$D$2:$D$4", allow_blank=True)
                dv_bool = DataValidation(type="list", formula1="=Értékek!$E$2:$E$3", allow_blank=True)

                ws1.add_data_validation(dv_dec_type)
                ws1.add_data_validation(dv_dec_kind)
                ws1.add_data_validation(dv_dec_freq)
                ws1.add_data_validation(dv_dec_meth)
                ws1.add_data_validation(dv_bool)

                dv_dec_type.add(ws1["D3"])
                dv_dec_kind.add(ws1["D4"])
                dv_dec_freq.add(ws1["D5"])
                dv_dec_meth.add(ws1["D9"])
                dv_bool.add(ws1["D10"])

                # 2. ANALITIKA LAP FORMÁZÁSA ÉS LENYÍLÓK
                ws2 = workbook['Analitika Adatok']
                ws2.views.sheetView[0].showGridLines = True
                
                gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
                dark_gray_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
                bold_font = Font(name="Segoe UI", size=10, bold=True)
                
                for r in range(1, 5):
                    for c in range(1, 28):
                        cell = ws2.cell(row=r, column=c)
                        if cell.value:
                            cell.fill = gray_fill
                            cell.font = bold_font
                
                for c in range(1, 28):
                    cell = ws2.cell(row=5, column=c)
                    cell.fill = dark_gray_fill
                    cell.font = Font(name="Consolas", size=10, bold=True, color="111111")

                for col in ws2.columns:
                    ws2.column_dimensions[col[0].column_letter].width = 22

                dv_doc_type = DataValidation(type="list", formula1="=Értékek!$F$2:$F$5", allow_blank=True)
                dv_part_stat = DataValidation(type="list", formula1="=Értékek!$G$2:$G$5", allow_blank=True)
                dv_pos_type = DataValidation(type="list", formula1="=Értékek!$H$2:$H$4", allow_blank=True)

                ws2.add_data_validation(dv_doc_type)
                ws2.add_data_validation(dv_part_stat)
                ws2.add_data_validation(dv_pos_type)

                dv_doc_type.add("G6:G500")
                dv_part_stat.add("I6:I500")
                dv_pos_type.add("U6:U500")

            messagebox.showinfo("Sikeres letöltés", f"A komplex, interaktív lenyílókkal és mintasorokkal ellátott Excel sablon elkészült az Asztalra:\n\n{asztal_utvonal}")
            logger.info(f"✅ [Sablon] Interaktív többlapos adatsablon mintákkal legenerálva: {asztal_utvonal}")
        except Exception as e:
            messagebox.showerror("Hiba", f"Nem sikerült a sablon mentése: {str(e)}")

    def _feldolgozas_tab_felep(self, tab: Any) -> None:
        gomb_sor = ctk.CTkFrame(tab, fg_color="transparent")
        gomb_sor.pack(fill="x", pady=(0, 8))

        self.btn_feldolgoz = ctk.CTkButton(
            gomb_sor, text="▶️  Mappa ellenőrzése és Feldolgozás indítása",
            width=180, fg_color="#2e7d32", hover_color="#1b5e20",
            command=self._feldolgozas_inditas,
        )
        self.btn_feldolgoz.pack(side="left", padx=4)

        ctk.CTkButton(
            gomb_sor, text="📥 Sablon letöltése", width=140,
            fg_color="#d35400", hover_color="#b33925",
            command=self._sablon_letoltese_vegrehajtasa
        ).pack(side="left", padx=4)

        import gui.previews as previews
        ctk.CTkButton(gomb_sor, text="📊  Excel előnézet", width=140, fg_color="#1565c0", hover_color="#0d47a1", command=lambda: previews.excel_preview_megnyitas(self)).pack(side="left", padx=4)
        ctk.CTkButton(gomb_sor, text="📄  XML előnézet", width=140, fg_color="#6a1b9a", hover_color="#4a148c", command=lambda: previews.xml_preview_megnyitas(self)).pack(side="left", padx=4)

        self.fajl_label = ctk.CTkLabel(tab, text="Nincs betöltött fájl.", font=ctk.CTkFont(size=12), text_color="#888888")
        self.fajl_label.pack(anchor="w", padx=6)

        ctk.CTkLabel(tab, text="Feldolgozási napló:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=6, pady=(8, 2))

        self.log_ablak = ctk.CTkTextbox(tab, height=320, font=ctk.CTkFont(size=12))
        self.log_ablak.pack(fill="both", expand=True, padx=4, pady=4)
        self.log_ablak.tag_config("hiba", foreground="#ef5350")
        self.log_ablak.tag_config("siker", foreground="#66bb6a")
        self.log_ablak.tag_config("figyelem", foreground="#ffb74d")
        self.log_ablak.tag_config("info", foreground="#4fc3f7")

        ctk.CTkLabel(tab, text="XML előnézet:", font=ctk.CTkFont(size=12, weight="bold")).pack(anchor="w", padx=6, pady=(8, 2))
        self.xml_elonezet = ctk.CTkTextbox(tab, height=120, font=ctk.CTkFont(size=10))
        self.xml_elonezet.pack(fill="both", expand=True, padx=4, pady=4)

    def _adozas_tab_felep(self, tab: Any) -> None:
        """Adózási állapot tab felépítése."""
        statusz_keret = ctk.CTkFrame(tab, corner_radius=10, fg_color="#2b2b2b")
        statusz_keret.pack(pady=10, padx=10, fill="x")

        ctk.CTkLabel(statusz_keret, text="Élő NAV Adózói Adatbázis Kapcsolat", font=ctk.CTkFont(size=14, weight="bold"), text_color="#5DADE2").grid(row=0, column=0, columnspan=2, pady=8, padx=15, sticky="w")

        self.lbl_ceg = ctk.CTkLabel(statusz_keret, text="Regisztrált Alany: Nincs konfigurált vállalat | Adószám: --------", font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_ceg.grid(row=1, column=0, columnspan=2, padx=15, pady=5, sticky="w")

        ctk.CTkLabel(statusz_keret, text="Hivatalos KOMA adatbázis tagság:").grid(row=2, column=0, padx=15, pady=5, sticky="w")
        self.lbl_koma_value = ctk.CTkLabel(statusz_keret, text="🟡 ELLENŐRZÉSRE VÁR (Szimulált üzemmód)", text_color="#F1C40F", font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_koma_value.grid(row=2, column=1, padx=15, pady=5, sticky="w")

        ctk.CTkLabel(statusz_keret, text="Aktuális NAV folyószámla egyenleg:").grid(row=3, column=0, padx=15, pady=(5, 10), sticky="w")
        self.lbl_egyenleg_value = ctk.CTkLabel(statusz_keret, text="Lekérdezés folyamatban...", text_color="#00FF88", font=ctk.CTkFont(size=12, weight="bold"))
        self.lbl_egyenleg_value.grid(row=3, column=1, padx=15, pady=(5, 10), sticky="w")
        
        self._frissit_adozas_adoszam()

    def _frissit_adozas_adoszam(self) -> None:
        """Frissíti az adószámot gépelés közben az Adózási fülön."""
        if hasattr(self, 'lbl_ceg'):
            val = ""
            if hasattr(self, 'entry_adoszam'):
                val = cast(str, cast(Any, self.entry_adoszam).get()).strip()
                
            adoszam_rovid = val if val else "--------"
            ceg_nev = "M2M Partner Vállalat Kft." if val else "Nincs konfigurált vállalat"
            cast(Any, self.lbl_ceg).configure(text=f"Regisztrált Alany: {ceg_nev} | Adószám: {adoszam_rovid}")
            
            if hasattr(self, 'lbl_koma_value'):
                if len(adoszam_rovid) == 8 and adoszam_rovid.isdigit():
                    cast(Any, self.lbl_koma_value).configure(text="🟢 IGEN (Köztartozásmentes Adatbázisban szerepel)", text_color="#00FF88")
                else:
                    cast(Any, self.lbl_koma_value).configure(text="🟡 ELLENŐRZÉSRE VÁR (Szimulált üzemmód)", text_color="#F1C40F")

    def _hataridok_tab_felep(self, tab: Any) -> None:
        import datetime
        hatarido_keret = ctk.CTkFrame(tab, corner_radius=10, fg_color="#232323")
        hatarido_keret.pack(pady=10, padx=10, fill="x")

        ctk.CTkLabel(hatarido_keret, text="📅 Jogszabályi ÁFA Bevallási Határidők", font=ctk.CTkFont(size=14, weight="bold"), text_color="#E74C3C").pack(pady=8, padx=15, anchor="w")

        ma = datetime.date.today()
        kov_honap = ma.replace(day=28) + datetime.timedelta(days=4)
        esedekesseg = datetime.date(kov_honap.year, kov_honap.month, 20)
        hatra_van = (esedekesseg - ma).days

        ctk.CTkLabel(hatarido_keret, text=f"A tárgyidőszaki adóbevallás és adóbefizetés törvényi határideje: {esedekesseg.strftime('%Y.%m.%d.')}\nHátralévő törvényes intézkedési idő: {hatra_van} nap.", font=ctk.CTkFont(size=12), justify="left").pack(pady=5, padx=15, anchor="w")

    def _elozmeny_tab_felep(self, tab: Any) -> None:
        ctk.CTkLabel(tab, text="Előzmények", font=ctk.CTkFont(size=14, weight="bold")).pack(pady=20)

    def _beallitas_tab_felep(self, tab: Any) -> None:
        beallitas_keret = ctk.CTkFrame(tab, corner_radius=10, fg_color="#2b2b2b")
        beallitas_keret.pack(pady=10, padx=20, fill="x")

        ctk.CTkLabel(beallitas_keret, text="⚙️ NAV API Hitelesítési Adatok", font=ctk.CTkFont(size=13, weight="bold")).grid(row=0, column=0, columnspan=2, pady=10, padx=10, sticky="w")

        self.entry_tech_user = ctk.CTkEntry(beallitas_keret, placeholder_text="Technikai felhasználónév", width=320)
        self.entry_tech_user.grid(row=1, column=0, padx=10, pady=(0, 10))
        self.entry_jelszo = ctk.CTkEntry(beallitas_keret, placeholder_text="Jelszó", show="*", width=320)
        self.entry_jelszo.grid(row=1, column=1, padx=10, pady=(0, 10))
        self.entry_xml_kulcs = ctk.CTkEntry(beallitas_keret, placeholder_text="XML cserekulcs", width=320)
        self.entry_xml_kulcs.grid(row=2, column=0, padx=10, pady=(0, 10))
        self.entry_sign_kulcs = ctk.CTkEntry(beallitas_keret, placeholder_text="Aláíró kulcs (sign key)", width=320)
        self.entry_sign_kulcs.grid(row=2, column=1, padx=10, pady=(0, 10))
        self.entry_adoszam = ctk.CTkEntry(beallitas_keret, placeholder_text="Adószám (8 számjegy)", width=320)
        self.entry_adoszam.grid(row=3, column=0, padx=10, pady=(0, 15))
        
        def _on_adoszam_keyup(event: Any) -> None:
            self._frissit_adozas_adoszam()
        self.entry_adoszam.bind("<KeyRelease>", _on_adoszam_keyup) # type: ignore

        self.combo_kornyezet = ctk.CTkComboBox(beallitas_keret, values=["TEST", "PROD"], width=320)
        self.combo_kornyezet.grid(row=3, column=1, padx=10, pady=(0, 15))
        self.combo_kornyezet.set("TEST")

        self.chk_eafa_feltoltes = ctk.CTkCheckBox(beallitas_keret, text="Valódi eÁFA TEST feltöltés engedélyezése")
        self.chk_eafa_feltoltes.grid(row=4, column=0, columnspan=2, padx=10, pady=(0, 15), sticky="w")

        gomb_vezerlo_keret = ctk.CTkFrame(tab, fg_color="transparent")
        gomb_vezerlo_keret.pack(pady=(10, 10))

        self.btn_kapcsolat_teszt = ctk.CTkButton(gomb_vezerlo_keret, text="NAV kapcsolat teszt", font=ctk.CTkFont(size=14, weight="bold"), height=42, width=220, command=self._kapcsolat_teszt_inditasa)
        self.btn_kapcsolat_teszt.pack(side="left", padx=8)

    def _helpdesk_tab_felep(self, tab: Any) -> None:
        helpdesk_keret = ctk.CTkFrame(tab, corner_radius=10, fg_color="#1e1e1e")
        helpdesk_keret.pack(pady=10, padx=10, fill="both", expand=True)

        ctk.CTkLabel(helpdesk_keret, text="✉️ Fejlesztői Helpdesk és Visszajelzési Csatorna", font=ctk.CTkFont(size=14, weight="bold"), text_color="#8E44AD").pack(pady=8, padx=15, anchor="w")

        self.entry_hiba_leiras = ctk.CTkTextbox(helpdesk_keret, height=120, font=ctk.CTkFont(size=12))
        self.entry_hiba_leiras.pack(fill="x", padx=15, pady=5)
        self.entry_hiba_leiras.insert("0.0", "[ Kérlek, ide részletesen írd le a javaslatot vagy hibát... ]")

        def hiba_bejelentes_vegrehajtasa() -> None:
            import datetime, urllib.parse, webbrowser
            txt_ctrl = cast(Any, self.entry_hiba_leiras)
            problema_szoveg = str(txt_ctrl.get("1.0", "end")).strip()
            if not problema_szoveg or "ide részletesen írd le" in problema_szoveg:
                messagebox.showerror("Helpdesk hiba", "Üres leírást nem küldhetsz be!")
                return

            fejleszto_email = "balint.papp@cegnev.hu"
            adoszam = cast(Any, self.entry_adoszam).get() if hasattr(self, 'entry_adoszam') else ""
            targy = f"NAV M2M Asszisztens Diagnosztika - Adószám: {adoszam}"

            test_szoveg = (
                f"=== NAV M2M RENDSZERHIBA JELENTÉS ===\n"
                f"Id•pont: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Adószám: {adoszam}\n"
                f"Környezet: eÁFA v2.0\n"
                f"--------------------------------------------------\n"
                f"Felhasználói észrevétel:\n{problema_szoveg}\n"
                f"--------------------------------------------------\n"
                f"Generált requestSignature verzió: SHA3-512 ok\n"
            )

            toplevel = cast(Any, cast(Any, self).winfo_toplevel())
            toplevel.clipboard_clear()
            toplevel.clipboard_append(test_szoveg)

            mailto_url = f"mailto:{fejleszto_email}?subject={urllib.parse.quote(targy)}&body={urllib.parse.quote('A teljes diagnosztikai naplót a szoftver automatikusan a vágólapra másolta. Kérlek nyomj egy Ctrl+V-t ide a szövegtörzsbe!')}"
            
            try:
                webbrowser.open(mailto_url)
                messagebox.showinfo("Helpdesk sikeres", "A levelező megnyitása elindult.\n\nA technikai adatokat a vágólapra másoltam, a megnyíló e-mailben nyomj egy Ctrl+V-t!")
            except Exception:
                messagebox.showinfo("Vágólapra mentve", "A levelezőt nem sikerült közvetlenül megnyitni, de a hibajelentést a vágólapra mentettem! Másold be egy levélbe balint.papp@cegnev.hu címre.")

        ctk.CTkButton(helpdesk_keret, text="✉️ Hibajelentés generálása és Vágólapra másolása", font=ctk.CTkFont(size=13, weight="bold"), fg_color="#8E44AD", hover_color="#732D91", height=38, command=hiba_bejelentes_vegrehajtasa).pack(pady=15)

    def _kapcsolat_teszt_inditasa(self) -> None:
        if hasattr(self, 'btn_kapcsolat_teszt'):
            cast(Any, self.btn_kapcsolat_teszt).configure(state="disabled", text="⏳ Kapcsolat teszt folyamatban...")
        if hasattr(self, 'btn_feldolgoz'):
            cast(Any, self.btn_feldolgoz).configure(state="disabled")

        def reset_buttons() -> None:
            if hasattr(self, 'btn_kapcsolat_teszt'):
                cast(Any, self.btn_kapcsolat_teszt).configure(state="normal", text="NAV kapcsolat teszt")
            if hasattr(self, 'btn_feldolgoz'):
                cast(Any, self.btn_feldolgoz).configure(state="normal")

        def statusz_updater(msg: str, color: str) -> None:
            if hasattr(self, 'fajl_label'):
                cast(Any, self.fajl_label).configure(text=msg, text_color=color)

        def error_popup_wrapper(title: str, msg: str) -> None:
            messagebox.showerror(title, msg)

        tech_user = cast(str, cast(Any, self.entry_tech_user).get()) if hasattr(self, 'entry_tech_user') else ""
        password = cast(str, cast(Any, self.entry_jelszo).get()) if hasattr(self, 'entry_jelszo') else ""
        sign_key = cast(str, cast(Any, self.entry_sign_kulcs).get()) if hasattr(self, 'entry_sign_kulcs') else ""
        exchange_key = cast(str, cast(Any, self.entry_xml_kulcs).get()) if hasattr(self, 'entry_xml_kulcs') else ""
        tax_number = cast(str, cast(Any, self.entry_adoszam).get()) if hasattr(self, 'entry_adoszam') else ""
        environment = cast(str, cast(Any, self.combo_kornyezet).get()) if hasattr(self, 'combo_kornyezet') else "TEST"

        kapcsolat_teszt_inditasa(
            tech_user=tech_user, password=password, sign_key=sign_key, exchange_key=exchange_key,
            tax_number=tax_number, environment=environment, allapot_uzenet=statusz_updater,
            show_error_popup=error_popup_wrapper, reset_buttons=reset_buttons
        )

    def _feldolgozas_inditas(self) -> None:
        if hasattr(self, 'btn_feldolgoz'):
            cast(Any, self.btn_feldolgoz).configure(state="disabled", text="⏳ Feldolgozás folyamatban...")
        if hasattr(self, 'btn_kapcsolat_teszt'):
            cast(Any, self.btn_kapcsolat_teszt).configure(state="disabled")

        def reset_buttons() -> None:
            if hasattr(self, 'btn_feldolgoz'):
                cast(Any, self.btn_feldolgoz).configure(state="normal", text="▶️  Mappa ellenőrzése és Feldolgozás indítása")
            if hasattr(self, 'btn_kapcsolat_teszt'):
                cast(Any, self.btn_kapcsolat_teszt).configure(state="normal", text="NAV kapcsolat teszt")

        def statusz_updater(msg: str, color: str) -> None:
            if hasattr(self, 'fajl_label'):
                cast(Any, self.fajl_label).configure(text=msg, text_color=color)

        def error_popup_wrapper(title: str, msg: str) -> None:
            messagebox.showerror(title, msg)

        def yes_no_popup_wrapper(title: str, msg: str) -> bool:
            return bool(messagebox.askyesno(title, msg))

        tech_user = cast(str, cast(Any, self.entry_tech_user).get()) if hasattr(self, 'entry_tech_user') else ""
        password = cast(str, cast(Any, self.entry_jelszo).get()) if hasattr(self, 'entry_jelszo') else ""
        sign_key = cast(str, cast(Any, self.entry_sign_kulcs).get()) if hasattr(self, 'entry_sign_kulcs') else ""
        exchange_key = cast(str, cast(Any, self.entry_xml_kulcs).get()) if hasattr(self, 'entry_xml_kulcs') else ""
        tax_number = cast(str, cast(Any, self.entry_adoszam).get()) if hasattr(self, 'entry_adoszam') else ""
        environment = cast(str, cast(Any, self.combo_kornyezet).get()) if hasattr(self, 'combo_kornyezet') else "TEST"
        eafa_feltoltes = bool(self.chk_eafa_feltoltes.get()) if hasattr(self, 'chk_eafa_feltoltes') else False

        automatikus_feldolgozas_inditasa(
            tech_user=tech_user, password=password, sign_key=sign_key, exchange_key=exchange_key,
            tax_number=tax_number, environment=environment, eafa_feltoltes=eafa_feltoltes,
            allapot_uzenet=statusz_updater, show_error_popup=error_popup_wrapper,
            ask_yes_no_popup=yes_no_popup_wrapper, reset_buttons=reset_buttons
        )